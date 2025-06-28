
CELLS_MAPPING = [
    ("ResCash", "G16:R16", "E-Cooking", "D2:O2"),

    ("ResCash", "G19:R19", "E-Cooking", "D4:O4"),
    ("ResCash", "G17:R17", "E-Cooking", "D5:O5"),
    ("ResCash", "G20:R20", "E-Cooking", "D6:O6"),
    ("ResCash", "G18:R18", "E-Cooking", "D7:O7"),
    ("Financial", "G13", "E-Cooking", "D8"),

    ("ResCash", "G23:R23", "E-Cooking", "D10:O10"),
    ("ResCash", "G21:R21", "E-Cooking", "D11:O11"),
    ("ResCash", "G24:R24", "E-Cooking", "D12:O12"),
    ("ResCash", "G22:R22", "E-Cooking", "D13:O13"),
    ("Financial", "G17", "E-Cooking", "D14"),

    ("ResCash", "G28:R28", "E-Cooking", "D16:O16"),
    ("ResCash", "G29:R29", "E-Cooking", "D17:O17"),
    ("ResCash", "G30:R30", "E-Cooking", "D18:O18"),
    ("ResCash", "G17", "E-Cooking", "D19"),

    ("ResCash", "G34:R34", "LPG", "D2:O2"),

    ("ResCash", "G35:R35", "LPG", "D4:O4"),
    ("ResCash", "G37:R37", "LPG", "D5:O5"),
    ("ResCash", "G36:R36", "LPG", "D6:O6"),
    ("ResCash", "G38:R38", "LPG", "D7:O7"),
    ("Financial", "G27", "LPG", "D8"),
    ("Financial", "G29", "LPG", "D9"),

    ("ResCash", "G40:R40", "LPG", "D11:O11"),
    ("ResCash", "G42:R42", "LPG", "D12:O12"),
    ("ResCash", "G41:R41", "LPG", "D13:O13"),
    ("ResCash", "G43:R43", "LPG", "D14:O14"),
    ("ResCash", "G39:R39", "LPG", "D15:O15"),
    ("Financial", "G32", "LPG", "D16"),

    ("ResCash", "G47:R58", "Rest of subsidies or taxes", "D2:O13"),
]



def read_range(ws, cell_range):
    if ':' in cell_range:
        cells = ws[cell_range]
        return [cell.value for cell in cells[0]]
    else:
        return [ws[cell_range].value]

def write_range(ws, start_cell, values):
    col_letters = ''.join(filter(str.isalpha, start_cell))
    row_number = int(''.join(filter(str.isdigit, start_cell)))

    from openpyxl.utils import column_index_from_string, get_column_letter
    start_col_index = column_index_from_string(col_letters)

    for i, val in enumerate(values):
        col_letter = get_column_letter(start_col_index + i)
        ws[f"{col_letter}{row_number}"].value = val

@app.post("/upload-technoeconomic-model/{name}")
def upload_techno_model(name: str, file: UploadFile = File(...)):
    global START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in VALID_EXTENSIONS:
        return {"error": "Only Excel files are allowed."}

    contents = file.file.read()
    wb_new = load_workbook(filename=BytesIO(contents), data_only=True)

    new_start = wb_new["Contents"]["G8"].value
    new_end = wb_new["Contents"]["G24"].value

    if START_YEAR_TECHNO_MODELS is not None and END_YEAR_TECHNO_MODELS is not None:
        if new_start != START_YEAR_TECHNO_MODELS or new_end != END_YEAR_TECHNO_MODELS:
            return {"error": f"Year range mismatch: existing {START_YEAR_TECHNO_MODELS}-{END_YEAR_TECHNO_MODELS}, new {new_start}-{new_end}. Upload cancelled."}
    else:
        START_YEAR_TECHNO_MODELS = new_start
        END_YEAR_TECHNO_MODELS = new_end
    
    FULL_TECHNO_INPUTS_PATH = f"{TECHNOECONOMIC_INPUTS_MODEL}{name}.xlsx"
    wb_base = load_workbook(FULL_TECHNO_INPUTS_PATH)

    for source_sheet, source_range, target_sheet, target_range in CELLS_MAPPING:
        ws_source = wb_new[source_sheet]
        ws_target = wb_base[target_sheet]

        values = read_range(ws_source, source_range)

        target_range = target_range if target_range else source_range

        write_range(ws_target, target_range.split(':')[0], values)

    wb_base.save(FULL_TECHNO_INPUTS_PATH)

    return {"status": "uploaded"}
