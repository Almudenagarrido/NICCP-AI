import os
import json
from itertools import product
from copy import deepcopy
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from copy import deepcopy

CELLS_MAPPING_MODEL_TO_INPUTS = [
    ("Res-Cash", "G16:R16", "Electricity", "D2:O2"),

    ("Res-Cash", "G19:R19", "Electricity", "D4:O4"),
    ("Res-Cash", "G17:R17", "Electricity", "D5:O5"),
    ("Res-Cash", "G20:R20", "Electricity", "D6:O6"),
    ("Res-Cash", "G18:R18", "Electricity", "D7:O7"),
    ("Financial", "G13", "Electricity", "D8"),

    ("Res-Cash", "G23:R23", "Electricity", "D10:O10"),
    ("Res-Cash", "G21:R21", "Electricity", "D11:O11"),
    ("Res-Cash", "G24:R24", "Electricity", "D12:O12"),
    ("Res-Cash", "G22:R22", "Electricity", "D13:O13"),
    ("Financial", "G17", "Electricity", "D14"),

    ("Res-Cash", "G28:R28", "Electricity", "D16:O16"),
    ("Res-Cash", "G29:R29", "Electricity", "D17:O17"),
    ("Res-Cash", "G30:R30", "Electricity", "D18:O18"),
    ("Financial", "G17", "Electricity", "D19"),

    ("Res-Cash", "G34:R34", "LPG", "D2:O2"),

    ("Res-Cash", "G35:R35", "LPG", "D4:O4"),
    ("Res-Cash", "G37:R37", "LPG", "D5:O5"),
    ("Res-Cash", "G36:R36", "LPG", "D6:O6"),
    ("Res-Cash", "G38:R38", "LPG", "D7:O7"),
    ("Financial", "G27", "LPG", "D8"),
    ("Financial", "G29", "LPG", "D9"),

    ("Res-Cash", "G40:R40", "LPG", "D11:O11"),
    ("Res-Cash", "G42:R42", "LPG", "D12:O12"),
    ("Res-Cash", "G41:R41", "LPG", "D13:O13"),
    ("Res-Cash", "G43:R43", "LPG", "D14:O14"),
    ("Res-Cash", "G39:R39", "LPG", "D15:O15"),
    ("Financial", "G32", "LPG", "D16"),

    ("Res-Cash", "G47:R58", "Rest of subsidies or taxes", "D2:O13"),
]

CELLS_MAPPING_MODEL_TO_CARBON = [
    ("Res-Cash", "G10:R10", "Carbon Credits", "C2:N2"),
]


OPS_MAP = {
    "addition": lambda x, y: x + y,
    "subtract": lambda x, y: x - y,
    "subtract_one": lambda x: x - 1,
    "multiply": lambda x, y: x * y,
    "multiply_per": lambda x, y: (x * y)/100,
    "divide": lambda x, y: x / y,
    "safe_divide": lambda x, y: x / y if y != 0 else 0,
    "safe_divide_subtract_one": lambda x, y: x / y - 1 if y != 0 else 0,
    "gt": lambda x, y: x > y,
    "gt_eq": lambda x, y: x > y,
    "lt": lambda x, y: x < y,
    "pos_or_cero": lambda x: max(x, 0),
    "equal": lambda x, y: x == y,
    "if": lambda condition, true_val, false_val: true_val if condition else false_val,
    "copy": lambda x: x,
    "negative": lambda x: -x,
    "percentage": lambda x: x*100,
    "min": lambda x, y: min(x, y),
    "max": lambda x, y: max(x, y),
    "abs": lambda x: abs(x),
}


def fill_contents_from_source(source_path: str, destiny_path: str, model_name:str, carbon_flag: bool):
    try:
        src_wb = load_workbook(source_path, data_only=True, read_only=True)
        dst_wb = load_workbook(destiny_path)

        cell_mapping = CELLS_MAPPING_MODEL_TO_INPUTS if not carbon_flag else CELLS_MAPPING_MODEL_TO_CARBON
        for src_sheet, src_range, dst_sheet, dst_range in cell_mapping:
            src_ws = src_wb[src_sheet]
            dst_ws = dst_wb[dst_sheet]

            src_min_col, src_min_row, src_max_col, src_max_row = range_boundaries(src_range)
            dst_min_col, dst_min_row, dst_max_col, dst_max_row = range_boundaries(dst_range)

            if carbon_flag:
                model_row = None
                for row in range(1, dst_ws.max_row + 1):
                    cell_value = dst_ws.cell(row=row, column=1).value
                    if isinstance(cell_value, str) and model_name.lower() in cell_value.strip().lower():
                        model_row = row
                        break
                if model_row is None:
                    raise ValueError(f"Model '{model_name}' not found in destination sheet '{dst_sheet}' column 1.")

                dst_min_row = model_row
                dst_max_row = model_row + (src_max_row - src_min_row)

            src_rows = src_max_row - src_min_row + 1
            src_cols = src_max_col - src_min_col + 1
            dst_rows = dst_max_row - dst_min_row + 1
            dst_cols = dst_max_col - dst_min_col + 1

            if src_rows != dst_rows or src_cols != dst_cols:
                raise ValueError(
                    f"Size mismatch between source {src_range} and destination {dst_range}."
                )

            for i in range(src_rows):
                for j in range(src_cols):
                    value = src_ws.cell(row=src_min_row + i, column=src_min_col + j).value
                    dst_ws.cell(row=dst_min_row + i, column=dst_min_col + j, value=value)

        dst_wb.save(destiny_path)

    except Exception as e:
        raise RuntimeError(f"Error actualizando plantilla con datos del modelo: {str(e)}")

def expand_formulas_by_models(formulas_json: dict, models: list, fuel_markets: list, ffss_sections: list) -> dict:
    expanded_json = {}

    for raw_file_path, sheets in formulas_json.items():
        file_paths = []

        if "{model}" in raw_file_path:
            file_paths = [raw_file_path.replace("{model}", m) for m in models]
        elif "{fuel_market}" in raw_file_path:
            file_paths = [raw_file_path.replace("{fuel_market}", f) for f in fuel_markets]
        elif "{ffss_section}" in raw_file_path:
            file_paths = [raw_file_path.replace("{ffss_section}", s) for s in ffss_sections]
        else:
            file_paths = [raw_file_path]

        for raw_sheet_name, formulas in sheets.items():
            for formula in formulas:
                targets = formula.get("targets", "")
                source_labels = formula.get("source_labels", [])

                contains_model = "{model}" in targets or any("{model}" in sl for sl in source_labels)
                contains_fuel = "{fuel_market}" in targets or any("{fuel_market}" in sl for sl in source_labels)
                contains_ffss = "{ffss_section}" in targets or any("{ffss_section}" in sl for sl in source_labels)

                for model, fuel, ffss in product(
                    models if contains_model else [None],
                    fuel_markets if contains_fuel else [None],
                    ffss_sections if contains_ffss else [None]
                ):
                    for file_path in file_paths:
                        new_formula = deepcopy(formula)

                        if contains_model and model is not None:
                            new_formula["targets"] = new_formula["targets"].replace("{model}", model)
                        if contains_fuel and fuel is not None:
                            new_formula["targets"] = new_formula["targets"].replace("{fuel_market}", fuel)
                        if contains_ffss and ffss is not None:
                            new_formula["targets"] = new_formula["targets"].replace("{ffss_section}", ffss)

                        new_labels = []
                        for sl in new_formula["source_labels"]:
                            if contains_model and model is not None:
                                sl = sl.replace("{model}", model)
                            if contains_fuel and fuel is not None:
                                sl = sl.replace("{fuel_market}", fuel)
                            if contains_ffss and ffss is not None:
                                sl = sl.replace("{ffss_section}", ffss)
                            new_labels.append(sl)

                        new_formula["source_labels"] = new_labels

                        try:
                            path_part, sheet_part, _ = new_formula["targets"].split("::")
                        except ValueError:
                            raise ValueError(f"Formato incorrecto en target: {new_formula['targets']}")

                        if path_part not in expanded_json:
                            expanded_json[path_part] = {}
                        if sheet_part not in expanded_json[path_part]:
                            expanded_json[path_part][sheet_part] = []

                        expanded_json[path_part][sheet_part].append(new_formula)

    return expanded_json

def parse_label_reference(default_file: str, default_sheet: str, label: str):
    parts = label.split("::")

    if len(parts) == 3:
        file_path, sheet, label_info = parts
    elif len(parts) == 2:
        file_path = default_file
        sheet, label_info = parts
    else:
        file_path = default_file
        sheet = default_sheet
        label_info = parts[0]

    if ":" in label_info:
        label_columns = label_info.split(":")
    else:
        label_columns = [label_info]

    return os.path.normpath(file_path), sheet, label_columns

def get_numeric_cell_refs_label(
    wbs: dict,
    default_file: str,
    default_sheet: str,
    full_label: str,
    first_label_col: int = 1
):
    file_path, sheet_name, label_columns = parse_label_reference(default_file, default_sheet, full_label)

    if file_path not in wbs:
        wbs[file_path] = load_workbook(file_path, data_only=True)
    ws = wbs[file_path][sheet_name]

    def find_best_matching_row():
        first_expected_str = label_columns[0].strip()
        candidate_rows = []

        for row in range(1, ws.max_row + 1):
            first_cell_value = ws.cell(row=row, column=first_label_col).value
            first_cell_str = str(first_cell_value).strip() if first_cell_value is not None else ""
            if first_cell_str == first_expected_str:
                candidate_rows.append(row)

        for row in candidate_rows:
            match = True
            for offset, expected_value in enumerate(label_columns[1:], start=1):
                col_index = first_label_col + offset
                cell_value = ws.cell(row=row, column=col_index).value
                cell_str = str(cell_value).strip() if cell_value is not None else ""
                expected_str = expected_value.strip()

                if cell_str != expected_str and expected_str not in cell_str:
                    match = False
                    break
            if match:
                return row

        for row in range(1, ws.max_row + 1):
            first_cell_value = ws.cell(row=row, column=first_label_col).value
            first_cell_str = str(first_cell_value).strip() if first_cell_value is not None else ""
            if first_expected_str not in first_cell_str:
                continue

            match = True
            for offset, expected_value in enumerate(label_columns[1:], start=1):
                col_index = first_label_col + offset
                cell_value = ws.cell(row=row, column=col_index).value
                cell_str = str(cell_value).strip() if cell_value is not None else ""
                expected_str = expected_value.strip()
                if expected_str not in cell_str:
                    match = False
                    break
            if match:
                return row

        return None


    target_row = find_best_matching_row()

    if target_row is None:
        raise ValueError(f"Label '{':'.join(label_columns)}' not found in sheet '{sheet_name}' starting at column {first_label_col}")

    refs = []
    for col in range(first_label_col + len(label_columns), ws.max_column + 1):
        val = ws.cell(row=target_row, column=col).value
        if isinstance(val, (int, float)):
            refs.append(ws.cell(row=target_row, column=col).coordinate)
        elif isinstance(val, str):
            try:
                float(val.replace(",", "."))
                refs.append(ws.cell(row=target_row, column=col).coordinate)
            except ValueError:
                pass

    return refs

def get_val(operand, values, results):
    if isinstance(operand, list):
        if operand[0] == "index":
            key = operand[1]
            if isinstance(key, int):
                return values[key] if key < len(values) else 0
            else:
                return results.get(key, 0)
        elif operand[0] == "literal":
            return operand[1]
    elif isinstance(operand, str):
        return results.get(operand, 0)
    elif isinstance(operand, (int, float)):
        return operand
    else:
        return 0
  
def get_cell_value_from_ref(wbs, ref, default_file, default_sheet):
        ref_parts = ref.split("::")
        if len(ref_parts) == 2:
            file_path = ref_parts[0]
            sheet_name, cell_coord = ref_parts[1].split("!")
        else:
            file_path = default_file
            sheet_name = default_sheet
            cell_coord = ref

        if file_path not in wbs:
            wbs[file_path] = load_workbook(file_path, data_only=True)

        return wbs[file_path][sheet_name][cell_coord]

def apply_formulas(file_path, formulas_json_path, models, fuel_markets, ffss_sections):

    with open(formulas_json_path) as f:
        formulas_json = json.load(f)
        formulas_json = expand_formulas_by_models(formulas_json, models, fuel_markets, ffss_sections)

        file_path_norm = os.path.normpath(file_path)
        json_keys_path = [os.path.normpath(k) for k in formulas_json.keys()]

        if file_path_norm not in json_keys_path:
            return
        
        json_index = json_keys_path.index(file_path_norm)
        formulas_for_file = list(formulas_json.values())[json_index]

        wbs = {}
        wbs[file_path_norm] = load_workbook(file_path_norm)
        
        for sheet_name, formulas in formulas_for_file.items():
            if sheet_name not in wbs[file_path_norm].sheetnames:
                continue
            ws = wbs[file_path_norm][sheet_name]
            
            for formula in formulas:
                targets_label = formula["targets"]
                source_labels = formula["source_labels"]
                formula_steps = formula.get("formula_steps", [])

                targets_cells = get_numeric_cell_refs_label(wbs, file_path_norm, sheet_name, targets_label, 1)
                source_cells_list = []
                for source_label in source_labels:
                    parts = source_label.split("::")
                    
                    if len(parts) == 3:
                        src_path, src_sheet, src_label = parts
                    else:
                        src_path, src_sheet = file_path_norm, sheet_name
                        src_label = source_label

                    src_path_norm = os.path.normpath(src_path)
                    if src_path_norm not in wbs:
                        wbs[src_path_norm] = load_workbook(src_path_norm)                 

                    refs = get_numeric_cell_refs_label(wbs, src_path_norm, src_sheet, src_label, 1)
                    source_cells_list.append((src_path_norm, src_sheet, refs))

                    length = len(targets_cells)
                    if any(len(refs) != length and len(refs) != 1 for (_, _, refs) in source_cells_list):
                        continue
                
                for i in range(length):
                    values = []
                    for src_path_norm, src_sheet, src_list in source_cells_list:
                        ref = src_list[i] if len(src_list) > 1 else src_list[0]
                        cell = get_cell_value_from_ref(wbs, ref, src_path_norm, src_sheet)
                        val = cell.value
                        if isinstance(val, str):
                            try:
                                val = float(val.replace(",", "."))
                            except:
                                val = 0
                        values.append(val if val is not None else 0)

                    results = {}
                    for step in formula_steps:
                        op = step["op"]
                        operands = step["operands"]
                        if op == "range":
                            res = i + 1
                        elif op == "offset":
                            source_index = operands[0][1]
                            offset = get_val(operands[1], values, results)
                
                            if 0 <= source_index < len(source_cells_list):
                                src_path_norm, src_sheet, src_list = source_cells_list[source_index]
                                j = i - offset
                                if 0 <= j < len(src_list):
                                    ref = src_list[j]
                                    cell = get_cell_value_from_ref(wbs, ref, src_path_norm, src_sheet)
                                    val = cell.value
                                    if isinstance(val, str):
                                        try:
                                            val = float(val.replace(",", "."))
                                        except:
                                            val = 0
                                    res = val if val is not None else 0
                                else:
                                    res = 0
                            else:
                                res = 0
                        elif op == "sum_range":
                            source_index = int(get_val(operands[0][1], values, results))
                            width = int(get_val(operands[1], values, results))
                            direction = step.get("direction", "backward")
                            fixed_start = step.get("start_index")
                            
                            if 0 <= source_index < len(source_cells_list):
                                src_path_norm, src_sheet, src_list = source_cells_list[source_index]
                                sum_vals = 0
                                if width == 0:
                                    width = len(src_list)
                                if fixed_start is not None:
                                    start = int(fixed_start)
                                    if direction == "forward":
                                        end = start + width
                                    else:
                                        start = start - width + 1
                                        end = start + width
                                else:
                                    if direction == "backward":
                                        start = i - width + 1
                                        end = i + 1
                                    elif direction == "forward":
                                        start = i
                                        end = i + width

                                for j in range(start, end):
                                    if 0 <= j < len(src_list):
                                        ref = src_list[j]
                                        cell = get_cell_value_from_ref(wbs, ref, src_path_norm, src_sheet)
                                        val = cell.value
                                        if isinstance(val, str):
                                            try:
                                                val = float(val.replace(",", "."))
                                            except:
                                                val = 0
                                        sum_vals += val if val is not None else 0
                                res = sum_vals
                            else:
                                res = 0
                        elif op == "first_nz_idx":
                            source_index = int(get_val(operands[0][1], values, results))

                            if 0 <= source_index < len(source_cells_list):
                                src_path_norm, src_sheet, src_list = source_cells_list[source_index]

                                row_vals = []
                                for ref in src_list:
                                    cell = get_cell_value_from_ref(wbs, ref, src_path_norm, src_sheet)
                                    val = cell.value
                                    if isinstance(val, str):
                                        try:
                                            val = float(val.replace(",", "."))
                                        except:
                                            val = 0
                                    row_vals.append(val if val is not None else 0)

                                res = next((idx for idx, v in enumerate(row_vals) if v != 0), 0)
                            else:
                                res = 0
                        elif op == "avg_range":
                            source_index = int(get_val(operands[0][1], values, results))
                            width = int(get_val(operands[1], values, results))
                            direction = step.get("direction", "backward")
                            fixed_start = step.get("start_index")

                            if 0 <= source_index < len(source_cells_list):
                                src_path_norm, src_sheet, src_list = source_cells_list[source_index]
                                values_list = []
                                if width == 0:
                                    width = len(src_list)
                                if fixed_start is not None:
                                    start = int(fixed_start)
                                    if direction == "forward":
                                        end = start + width
                                    else:
                                        start = start - width + 1
                                        end = start + width
                                else:
                                    if direction == "backward":
                                        start = i - width + 1
                                        end = i + 1
                                    elif direction == "forward":
                                        start = i
                                        end = i + width

                                for j in range(start, end):
                                    if 0 <= j < len(src_list):
                                        ref = src_list[j]
                                        cell = get_cell_value_from_ref(wbs, ref, src_path_norm, src_sheet)
                                        val = cell.value
                                        if isinstance(val, str):
                                            try:
                                                val = float(val.replace(",", "."))
                                            except:
                                                val = 0
                                        values_list.append(val if val is not None else 0)

                                res = sum(values_list) / len(values_list) if values_list else 0
                            else:
                                res = 0
                        else:
                            ops_args = [get_val(operand, values, results) for operand in operands]
                            if op in OPS_MAP:
                                res = OPS_MAP[op](*ops_args)
                            else:
                                print(f"Operation {op} not supported by platform")
                                res = 0
                                
                        results[step["result"]] = res

                    target_ref = targets_cells[i]
                    ws[target_ref].value = results.get("final", 0)

        wbs[file_path_norm].save(file_path_norm)


"""
----- PROBLEMILLAS -----
Comprobar flujo de las entradas upstream, local, grid, off-grid.
Calculo de Long term subisdies a partir de entradas existentes.
Paises precargados, permitir varios modelos para cada pais.
Reestructuracion de backend (optimizacion de endpoints y clases).
Gestion de plantilla y formato para guardar datos (json, excel, etc).
EBITDA margin y OPEX subsidies en LPG Financial Inputs (Execution).

Relacionado con los SGD 2030, definir rango de años como entrada. Afecta a:
- + Increase en GRANTS en CS (valores para Baseline y 2024, se conceden en dos años); Realisation (% capex) (suma desde 2023 hasta 2030)

---------------------------------------------
VALORES VACIOS
---------------------------------------------
- Balance Sheet: Intangibles, Other fixed assets, Other financial assets, Other current tax assets, Share Premium, Deferred tax liabilities, Other short term liabilities, Other tax liabilities, Short Term financial liabilities (depende de D163-O163 que son celdas vacias), Provisions

- PP&E - Capex: Divestiture

- Working Capital: Inventories - Days of COGS 

- En la hoja de CAPEX para Electricity (o LPG) Maintenance CAPEX y Acc. Growth CAPEX.
"""