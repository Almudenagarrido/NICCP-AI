from fastapi import FastAPI, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.responses import FileResponse
import os
import re
from openpyxl import load_workbook
from io import BytesIO
import shutil
import tempfile
import pandas as pd
from pydantic import BaseModel
from typing import List, Dict
import excel_utils as e


FUEL_MARKET_INFORMATION_FOLDER = "./fuel-market-information"
FUEL_MARKET_INFORMATION_PATH = os.path.join(FUEL_MARKET_INFORMATION_FOLDER, "fuel-market-information.xlsx")
FUEL_MARKET_INFORMATION_TEMPLATE = os.path.join(FUEL_MARKET_INFORMATION_FOLDER, "fuel-market-information-template.xlsx")
TECHNOECONOMIC_MODELS_FOLDER = "uploaded-technoeconomic-models"
DESIGN_CAPITAL_STRUCTURE_FOLDER = "./design-capital-structure"
DESIGN_CAPITAL_STRUCTURE_TEMPLATE = os.path.join(DESIGN_CAPITAL_STRUCTURE_FOLDER, "design-capital-structure-template.xlsx")
DESIGN_CAPITAL_STRUCTURE_MODEL = os.path.join(DESIGN_CAPITAL_STRUCTURE_FOLDER, "design-capital-structure-")
TECHNOECONOMIC_INPUTS_FOLDER = "./technoeconomic-inputs"
TECHNOECONOMIC_INPUTS_TEMPLATE = os.path.join(TECHNOECONOMIC_INPUTS_FOLDER, "technoeconomic-inputs-template.xlsx")
TECHNOECONOMIC_INPUTS_MODEL = os.path.join(TECHNOECONOMIC_INPUTS_FOLDER, "technoeconomic-inputs-")
CARBON_CREDITS_FOLDER = "./carbon-credits"
CARBON_CREDITS_TEMPLATE_PATH = os.path.join(CARBON_CREDITS_FOLDER, "carbon-credits-template.xlsx")
CARBON_CREDITS_PATH = os.path.join(CARBON_CREDITS_FOLDER, "carbon-credits.xlsx")
FINANCIAL_STATEMENTS_FOLDER = "./financial-statements"
FINANCIAL_STATEMENTS_TEMPLATE = os.path.join(FINANCIAL_STATEMENTS_FOLDER, "financial-statements-template.xlsx")
FINANCIAL_STATEMENTS_MODEL = os.path.join(FINANCIAL_STATEMENTS_FOLDER, "financial-statements-")
CAPEX_MARKET_FOLDER = "./capex-fuel-markets"
CAPEX_MARKET_TEMPLATE = os.path.join(CAPEX_MARKET_FOLDER, "capex-market-template.xlsx")
CAPEX_MARKET_MODEL = os.path.join(CAPEX_MARKET_FOLDER, "capex-market-")
FORMULAS_JSON_PATH = "./formulas_map.json"
FOLDERS = [
    TECHNOECONOMIC_MODELS_FOLDER,
    TECHNOECONOMIC_INPUTS_FOLDER,
    DESIGN_CAPITAL_STRUCTURE_FOLDER,
    CARBON_CREDITS_FOLDER, 
    FUEL_MARKET_INFORMATION_FOLDER,
    FINANCIAL_STATEMENTS_FOLDER,
    CAPEX_MARKET_FOLDER,
]

VALID_EXTENSIONS = (".xlsx", ".xlsm", ".xls", ".xltx", ".xltm")
START_YEAR_TECHNO_MODELS = None
END_YEAR_TECHNO_MODELS = None


class SheetUpdate(BaseModel):
    model:str
    sheet_name: str
    data: List[Dict] 

class MarketName(BaseModel):
    name: str

app = FastAPI()


# Fuel Market Information
@app.get("/fuel-market-information")
def get_fuel_market_information():
    global START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS
    
    models = [
        os.path.splitext(f)[0].removeprefix("technoeconomic-inputs-")
        for f in os.listdir(TECHNOECONOMIC_INPUTS_FOLDER)
        if f.lower().endswith(tuple(ext.lower() for ext in VALID_EXTENSIONS)) and "template" not in f.lower()
    ]

    if not os.path.exists(FUEL_MARKET_INFORMATION_PATH):
        if os.path.exists(FUEL_MARKET_INFORMATION_TEMPLATE):
            shutil.copy(FUEL_MARKET_INFORMATION_TEMPLATE, FUEL_MARKET_INFORMATION_PATH)
        else:
            return {"error": "Template file is missing"}
    
    wb = load_workbook(FUEL_MARKET_INFORMATION_PATH)

    for ws in wb.worksheets:
        year_row = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, str) and cell_val.strip().lower() == "baseline":
                    year_row = row
                    break
            if year_row:
                break
        
        if year_row is None:
            continue

        cols_to_delete = []
        for col in range(3, ws.max_column + 1):
            cell_val = ws.cell(row=year_row, column=col).value
            try:
                year = int(cell_val)
                if year <= START_YEAR_TECHNO_MODELS or year > END_YEAR_TECHNO_MODELS:
                    cols_to_delete.append(col)
            except (TypeError, ValueError):
                pass

        for col_idx in reversed(cols_to_delete):
            ws.delete_cols(col_idx)

        if ws.title == "LPG":
            existing_rows = {}

            for i in range(2, ws.max_row + 1):
                val = ws.cell(row=i, column=1).value
                if val:
                    key = str(val).strip()
                    existing_rows[key] = i
            
            expected_rows = {
                f"% EBITDA Margin - {model}" for model in models
            }.union({
                f"OPEX subsidies - {model}" for model in models
            })

            for row_label, row_num in list(existing_rows.items())[::-1]:
                if (row_label.startswith("% EBITDA Margin - ") or row_label.startswith("OPEX subsidies - ")) and row_label not in expected_rows:
                    ws.delete_rows(row_num)

            existing_inputs = {
                str(ws.cell(row=i, column=1).value).strip()
                for i in range(2, ws.max_row + 1)
                if ws.cell(row=i, column=1).value
            }

            for model_name in models:
                for row_label in [f"% EBITDA Margin - {model_name}", f"OPEX subsidies - {model_name}"]:
                    if row_label not in existing_inputs:
                        new_row = ws.max_row + 1
                        ws.cell(row=new_row, column=1).value = row_label
                        ws.cell(row=new_row, column=2).value = "%"
                        for col in range(3, ws.max_column + 1):
                            ws.cell(row=new_row, column=col).value = 0

    wb.save(FUEL_MARKET_INFORMATION_PATH)

    return FileResponse(
        FUEL_MARKET_INFORMATION_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="fuel-market-information.xlsx"
    )

@app.post("/save-fuel-market-information")
async def save_fuel_market_information(update: SheetUpdate):

    df_new = pd.DataFrame(update.data)

    if 'index' in df_new.columns:
        df_new.set_index('index', inplace=True)

    excel_path = FUEL_MARKET_INFORMATION_PATH
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_new.to_excel(writer, sheet_name=update.sheet_name, index=False)

    return {"message": "The file with the fuel market information was updated successfully"}

@app.post("/reset-fuel-market-information")
def reset_fuel_market_information(update: SheetUpdate):
    
    if not os.path.exists(FUEL_MARKET_INFORMATION_PATH):
        return {"error": "Fuel market information file does not exist"}
    
    if not os.path.exists(FUEL_MARKET_INFORMATION_TEMPLATE):
        return {"error": "Template file is missing"}

    try:
        if update.sheet_name not in {"Carbon", "LPG"}:
            template_sheet = "Electricity"
        else:
            template_sheet = update.sheet_name

        template_df = pd.read_excel(FUEL_MARKET_INFORMATION_TEMPLATE, sheet_name=template_sheet)
        wb = load_workbook(FUEL_MARKET_INFORMATION_PATH)

        if update.sheet_name in wb.sheetnames:
            std = wb[update.sheet_name]
            wb.remove(std)

        wb.save(FUEL_MARKET_INFORMATION_PATH)

        with pd.ExcelWriter(FUEL_MARKET_INFORMATION_PATH, engine='openpyxl', mode='a') as writer:
            template_df.to_excel(writer, sheet_name=update.sheet_name, index=False)

        return {"message": f"'{update.sheet_name}' was reset to template version"}

    except Exception as e:
        return {"error": f"Failed to reset: {str(e)}"}

@app.post("/add-fuel-market")
def add_fuel_market(market: MarketName):
    if not os.path.exists(FUEL_MARKET_INFORMATION_PATH):
        return {"error": "fuel-market-information.xlsx not found"}

    xls_template = pd.ExcelFile(FUEL_MARKET_INFORMATION_TEMPLATE, engine="openpyxl")
    xls = pd.ExcelFile(FUEL_MARKET_INFORMATION_PATH, engine="openpyxl")
    df_ecooking = pd.read_excel(xls_template, sheet_name="Electricity", engine="openpyxl")
    new_sheet_name = market.name
    
    if "Electricity" not in xls_template.sheet_names:
        return {"error": "Base sheet 'Electricity' not found"}

    if new_sheet_name in xls.sheet_names:
        return {"error": f"The sheet '{new_sheet_name}' already exists."}

    sheets_dict = {sheet: xls.parse(sheet) for sheet in xls.sheet_names}
    sheets_dict[new_sheet_name] = df_ecooking.copy()

    with pd.ExcelWriter(FUEL_MARKET_INFORMATION_PATH, engine="openpyxl", mode="w") as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    return {"message": f"Market '{new_sheet_name}' added successfully."}

@app.post("/delete-fuel-market")
def delete_fuel_market(market: MarketName):
    if not os.path.exists(FUEL_MARKET_INFORMATION_PATH):
        return {"error": "fuel-market-information.xlsx not found"}

    xls = pd.ExcelFile(FUEL_MARKET_INFORMATION_PATH, engine="openpyxl")
    sheet_to_delete = market.name

    if sheet_to_delete not in xls.sheet_names:
        return {"error": f"The sheet '{sheet_to_delete}' does not exist."}

    remaining_sheets = {
        sheet: xls.parse(sheet)
        for sheet in xls.sheet_names if sheet != sheet_to_delete
    }

    with pd.ExcelWriter(FUEL_MARKET_INFORMATION_PATH, engine="openpyxl", mode="w") as writer:
        for name, df in remaining_sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

    return {"message": f"Market '{sheet_to_delete}' deleted successfully."}


# Techno-Economic Models
@app.get("/technoeconomic-models")
def list_techno_models():
    if not os.path.exists(CARBON_CREDITS_PATH):
        return {"models": []}

    wb = load_workbook(CARBON_CREDITS_PATH)
    ws = wb["Carbon Credits"]

    model_names = set()
    pattern = r"CO2 emited\s*-\s*(.+?)\s*scenario"

    for _, row in enumerate(ws.iter_rows(), start=1):
        cell = row[0].value if len(row) > 0 else None
        if isinstance(cell, str):
            cell_cleaned = re.sub(r"\{.*?\}", "", cell)
            match = re.search(pattern, cell_cleaned)
            if match:
                model = match.group(1).strip()
                if model:
                    model_names.add(model)
    
    models_sorted = sorted(model_names)
    if "BAU" in models_sorted:
        models_sorted.remove("BAU")
        models_sorted.insert(0, "BAU")

    return {"models": models_sorted}

def detect_year_range(path):
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                if isinstance(ws.cell(row=row, column=col).value, str) and ws.cell(row=row, column=col).value.strip().lower() == "baseline":
                    years = []
                    for c in range(col + 1, ws.max_column + 1):
                        try:
                            y = int(ws.cell(row=row, column=c).value)
                            years.append(y)
                        except Exception:
                            continue
                    if years:
                        return min(years) - 1, max(years)
    return None, None

@app.post("/create-technoeconomic-model/{model}")
async def create_technoeconomic_model(model: str, start_year: int, end_year: int):
    global START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS

    model = model.strip()
    if not model:
        return {"error": "Model name cannot be empty"}

    if start_year >= end_year:
        return {"error": "Start year must be less than end year"}

    if model.lower() == "bau":
        file_path = CARBON_CREDITS_PATH
        template_path = CARBON_CREDITS_TEMPLATE_PATH
    else:
        file_path = f"{TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx"
        template_path = TECHNOECONOMIC_INPUTS_TEMPLATE

    if not os.path.exists(file_path):
        if not os.path.exists(template_path):
            return {"error": f"Template file missing: {template_path}"}
        shutil.copy(template_path, file_path)

    if model != "bau":
        wb_cc = load_workbook(CARBON_CREDITS_PATH)
        ws_cc = wb_cc["Carbon Credits"]

        rows_to_insert = []
        for row_idx in range(1, ws_cc.max_row + 1):
            cell_val = ws_cc.cell(row=row_idx, column=1).value
            if cell_val and isinstance(cell_val, str) and "{model}" in cell_val:
                row_values = [ws_cc.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws_cc.max_column + 1)]
                rows_to_insert.append((row_idx, row_values))

        for orig_row_idx, row_values in reversed(rows_to_insert):
            insert_at = orig_row_idx + 1
            ws_cc.insert_rows(insert_at)

            for col_idx, val in enumerate(row_values, start=1):
                if col_idx == 1 and isinstance(val, str):
                    val = val.replace("{model}", model)
                ws_cc.cell(row=insert_at, column=col_idx, value=val)

        wb_cc.save(CARBON_CREDITS_PATH)

        ref_start, ref_end = detect_year_range(CARBON_CREDITS_PATH)
        if ref_start != start_year or ref_end != end_year:
            return {
                "error": f"Year range mismatch: expected {ref_start}-{ref_end}, got {start_year}-{end_year}"
            }
        
    if START_YEAR_TECHNO_MODELS is None or END_YEAR_TECHNO_MODELS is None:
        START_YEAR_TECHNO_MODELS = start_year
        END_YEAR_TECHNO_MODELS = end_year
        
    wb = load_workbook(file_path)

    if model == "bau":
        ws = wb["Carbon Credits"]
        year_row = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, str) and cell_val.strip().lower() == "baseline":
                    year_row = row
                    break
            if year_row:
                break
            
        if year_row:
            cols_to_delete = []
            for col in range(3, ws.max_column + 1):
                cell_val = ws.cell(row=year_row, column=col).value
                try:
                    year = int(cell_val)
                    if year <= START_YEAR_TECHNO_MODELS or year > END_YEAR_TECHNO_MODELS:
                        cols_to_delete.append(col)
                except (TypeError, ValueError):
                    continue
            
            for col_idx in reversed(cols_to_delete):
                ws.delete_cols(col_idx)
    else:
        xls_fuel_market = pd.ExcelFile(FUEL_MARKET_INFORMATION_PATH, engine="openpyxl")
        fuel_market_sheets = xls_fuel_market.sheet_names
        expected_sheets = [
            "C02" if name.strip().lower() == "carbon" else name
            for name in fuel_market_sheets
        ]

        xls_model = pd.ExcelFile(file_path, engine="openpyxl")
        existing_sheets = xls_model.sheet_names

        xls_template = pd.ExcelFile(template_path, engine="openpyxl")
        df_template = pd.read_excel(xls_template, sheet_name="LPG", engine="openpyxl")

        missing_sheets = [s for s in expected_sheets if s not in existing_sheets]

        if missing_sheets:
            sheets_dict = {
                sheet: pd.read_excel(xls_model, sheet_name=sheet)
                for sheet in existing_sheets
            }
            for missing in missing_sheets:
                sheets_dict[missing] = df_template.copy()
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                for sheet_name, df in sheets_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        sheets_to_remove = [
            ws.title for ws in wb.worksheets
            if ws.title not in expected_sheets and ws.title != "Rest of subsidies or taxes"
        ]
        for sheet_name in sheets_to_remove:
            wb.remove(wb[sheet_name])

        for ws in wb.worksheets:
            year_row = None
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=row, column=col).value
                    if isinstance(cell_val, str) and cell_val.strip().lower() == "baseline":
                        year_row = row
                        break
                if year_row:
                    break

            if year_row is None:
                continue

            cols_to_delete = []
            for col in range(3, ws.max_column + 1):
                cell_val = ws.cell(row=year_row, column=col).value
                try:
                    year = int(cell_val)
                    if year <= START_YEAR_TECHNO_MODELS or year > END_YEAR_TECHNO_MODELS:
                        cols_to_delete.append(col)
                except (TypeError, ValueError):
                    continue

            for col_idx in reversed(cols_to_delete):
                ws.delete_cols(col_idx)

    wb.save(file_path)
    wb.close()

    return {"success": True, "message": f"Model '{model}' created successfully."}

@app.get("/download-technoeconomic-model-files/{model}")
def download_techno_model_files(model: str, background_tasks: BackgroundTasks):
    temp_dir = tempfile.mkdtemp()
    try:
        paths = [
            FUEL_MARKET_INFORMATION_PATH,
            f"{TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx",
            f"{DESIGN_CAPITAL_STRUCTURE_MODEL}{model}.xlsx",
        ]

        found_model_file = None
        for f in os.listdir(TECHNOECONOMIC_MODELS_FOLDER):
            if model.lower() in f.lower():
                candidate_path = os.path.join(TECHNOECONOMIC_MODELS_FOLDER, f)
                if os.path.isfile(candidate_path):
                    found_model_file = candidate_path
                    break

        if found_model_file:
            paths.append(found_model_file)

        for path in paths:
            if not os.path.exists(path):
                shutil.rmtree(temp_dir)
                return {"error": f"Missing file: {path}"}
            shutil.copy(path, os.path.join(temp_dir, os.path.basename(path)))

        zip_path = os.path.join(tempfile.gettempdir(), f"{model}_technoeconomic_files.zip")
        shutil.make_archive(zip_path.replace(".zip", ""), 'zip', temp_dir)

        background_tasks.add_task(os.remove, zip_path)
        shutil.rmtree(temp_dir)

        return FileResponse(
            zip_path,
            media_type="application/zip",
            filename=f"{model}_technoeconomic_files.zip"
        )
    except Exception as e:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        return {"error": str(e)}

@app.delete("/delete-technoeconomic-model/{model}")
def delete_techno_model(model: str):
    errors = []
    model_lower = model.lower()

    if model_lower == "bau":
        for folder in FOLDERS:
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                if os.path.isfile(file_path) and "template" not in filename.lower():
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        errors.append(f"Failed to delete {file_path}: {str(e)}")
    else:
        paths_to_delete = [
            os.path.join(TECHNOECONOMIC_MODELS_FOLDER, f)
            for f in os.listdir(TECHNOECONOMIC_MODELS_FOLDER)
            if model_lower in f.lower()
        ]

        related_files = [
            f"{TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx",
            f"{DESIGN_CAPITAL_STRUCTURE_MODEL}{model}.xlsx",
        ]

        for path in related_files:
            paths_to_delete.append(path)

        for path in paths_to_delete:
            try:
                if os.path.exists(path) and os.path.abspath(path) != os.path.abspath(FUEL_MARKET_INFORMATION_PATH):
                    os.remove(path)
            except Exception as e:
                errors.append(f"Failed to delete {path}: {str(e)}")

    try:
        carbon_file = f"{CARBON_CREDITS_PATH}"
        sheet = "Carbon Credits"

        df = pd.read_excel(carbon_file, sheet_name=sheet)
        first_col = df.columns[0]
        original_col = df[first_col]

        mask = ~original_col.astype(str).str.lower().str.contains(model.lower())
        df_filtered = df[mask]

        with pd.ExcelWriter(carbon_file, mode='a', if_sheet_exists='replace') as writer:
            df_filtered.to_excel(writer, sheet_name=sheet, index=False)
    except Exception as e:
        errors.append(f"Failed to update Carbon Credits file: {str(e)}")
    if errors:
        raise HTTPException(status_code=500, detail={"errors": errors})

    return {"status": "deleted"}

@app.post("/upload-technoeconomic-model/{name}")
def upload_techno_model(name: str, file: UploadFile = File(...)):
    global START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS

    ext = os.path.splitext(file.filename)[1].lower()
    
    if ext not in VALID_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only Excel files are allowed.")

    contents = file.file.read()
    new_file = load_workbook(filename=BytesIO(contents), data_only=True, read_only=True)

    new_start = new_file["Contents"]["G8"].value
    new_end = new_file["Contents"]["G24"].value

    if name.lower() == "bau":
        reference_path = CARBON_CREDITS_PATH
        if not os.path.exists(reference_path):
            shutil.copy(CARBON_CREDITS_TEMPLATE_PATH, reference_path)
    else:
        reference_path = f"{TECHNOECONOMIC_INPUTS_MODEL}{name}.xlsx"
        if not os.path.exists(reference_path):
            raise HTTPException(status_code=400, detail=f"Reference technoeconomic model not found: {reference_path}")

    ref_start, ref_end = detect_year_range(reference_path)
    if (new_start != ref_start) or (new_end != ref_end):
        raise HTTPException(
            status_code=400,
            detail=f"Year range mismatch: expected {ref_start}-{ref_end}, got {new_start}-{new_end}"
        )

    UPLOADED_MODEL_PATH = os.path.join(TECHNOECONOMIC_MODELS_FOLDER, name + ext)
    with open(UPLOADED_MODEL_PATH, "wb") as f:
            f.write(contents)

    e.fill_contents_from_source(UPLOADED_MODEL_PATH, CARBON_CREDITS_PATH, name, carbon_flag=True)

    if name.lower() != "bau":
        e.fill_contents_from_source(UPLOADED_MODEL_PATH, reference_path, name, carbon_flag=False)

    return {"status": "uploaded"}


# Design Capital Structure
@app.get("/design-capital-structure/{model}")
async def get_design_capital_structure(model: str):
    global START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS
    if model.lower() in ["none", "bau"]:
        if not os.path.exists(DESIGN_CAPITAL_STRUCTURE_TEMPLATE):
            return {"error": "Template file is missing"}
        
        return FileResponse(
            DESIGN_CAPITAL_STRUCTURE_TEMPLATE,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=DESIGN_CAPITAL_STRUCTURE_TEMPLATE
        )

    FULL_DESIGN_MODEL_PATH = f"{DESIGN_CAPITAL_STRUCTURE_MODEL}{model}.xlsx"

    if not os.path.exists(FUEL_MARKET_INFORMATION_PATH):
        return {"error": "fuel-market-information.xlsx not found"}

    if not os.path.exists(DESIGN_CAPITAL_STRUCTURE_TEMPLATE):
        return {"error": "Template file is missing"}

    if not os.path.exists(FULL_DESIGN_MODEL_PATH):
        shutil.copy(DESIGN_CAPITAL_STRUCTURE_TEMPLATE, FULL_DESIGN_MODEL_PATH)

    if START_YEAR_TECHNO_MODELS is None and END_YEAR_TECHNO_MODELS is None:
        path = f"{TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx"
        START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS = detect_year_range(path)

    xls_fuel_market = pd.ExcelFile(FUEL_MARKET_INFORMATION_PATH, engine="openpyxl")
    fuel_market_sheets = xls_fuel_market.sheet_names
    expected_sheets = [
        "Electricity (Low access)" if name.strip().lower() == "carbon"
        else "Electricity & E-Cooking" if name.strip().lower() == "electricity"
        else name
        for name in fuel_market_sheets
    ]

    xls_model = pd.ExcelFile(FULL_DESIGN_MODEL_PATH, engine="openpyxl")
    existing_sheets = xls_model.sheet_names

    xls_template = pd.ExcelFile(DESIGN_CAPITAL_STRUCTURE_TEMPLATE, engine="openpyxl")
    df_ecooking = pd.read_excel(xls_template, sheet_name="Electricity & E-Cooking", engine="openpyxl", header=None)

    missing_sheets = [s for s in expected_sheets if s not in existing_sheets]

    if missing_sheets:
        sheets_dict = {
            sheet: pd.read_excel(xls_model, sheet_name=sheet)
            for sheet in existing_sheets
        }
        for missing in missing_sheets:
            sheets_dict[missing] = df_ecooking.copy()
        with pd.ExcelWriter(FULL_DESIGN_MODEL_PATH, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(FULL_DESIGN_MODEL_PATH)

    sheets_to_remove = [ws.title for ws in wb.worksheets if ws.title not in expected_sheets]

    for sheet_name in sheets_to_remove:
        std = wb[sheet_name]
        wb.remove(std)

    for ws in wb.worksheets:
        year_row = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, str) and cell_val.strip().lower() == "baseline":
                    year_row = row
                    break

            if year_row:
                break
        
        if year_row is None:
            continue

        cols_to_delete = []
        for col in range(3, ws.max_column + 1):
            cell_val = ws.cell(row=year_row, column=col).value
            try:
                year = int(cell_val)
                if year <= START_YEAR_TECHNO_MODELS or year > END_YEAR_TECHNO_MODELS:
                    cols_to_delete.append(col)
            except (TypeError, ValueError):
                pass

        for col_idx in reversed(cols_to_delete):
            ws.delete_cols(col_idx)

    wb.save(FULL_DESIGN_MODEL_PATH)

    return FileResponse(
        FULL_DESIGN_MODEL_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=FULL_DESIGN_MODEL_PATH
    )

@app.post("/save-design-capital-structure")
async def save_design_capital_structure(update: SheetUpdate):
    FULL_DESIGN_MODEL_PATH = f"{DESIGN_CAPITAL_STRUCTURE_MODEL}{update.model}.xlsx"

    if not os.path.exists(FULL_DESIGN_MODEL_PATH):
        return {"error": f"File for model '{update.model}' not found"}

    try:
        df = pd.DataFrame(update.data)
        if 'index' in df.columns:
            df.set_index('index', inplace=True)

        with pd.ExcelWriter(FULL_DESIGN_MODEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=update.sheet_name, index=False)

        return {"message": f"Design capital structure for model '{update.model}' saved successfully."}
    except Exception as e:
        return {"error": f"Failed to save: {str(e)}"}

@app.post("/reset-design-capital-structure")
def reset_design_capital_structure(update: SheetUpdate):
    FULL_DESIGN_MODEL_PATH = f"{DESIGN_CAPITAL_STRUCTURE_MODEL}{update.model}.xlsx"

    if not os.path.exists(FULL_DESIGN_MODEL_PATH):
        return {"error": f"Model file '{FULL_DESIGN_MODEL_PATH}' not found"}
    
    if not os.path.exists(DESIGN_CAPITAL_STRUCTURE_TEMPLATE):
        return {"error": f"Template file '{DESIGN_CAPITAL_STRUCTURE_TEMPLATE}' not found"}

    try:
        if update.sheet_name not in {"Electricity", "LPG"}:
            template_sheet = "Electricity & E-Cooking"
        else:
            template_sheet = update.sheet_name

        template_df = pd.read_excel(DESIGN_CAPITAL_STRUCTURE_TEMPLATE, sheet_name=template_sheet)
        wb = load_workbook(FULL_DESIGN_MODEL_PATH)

        if update.sheet_name in wb.sheetnames:
            std = wb[update.sheet_name]
            wb.remove(std)
            wb.save(FULL_DESIGN_MODEL_PATH)

        with pd.ExcelWriter(FULL_DESIGN_MODEL_PATH, engine='openpyxl', mode='a') as writer:
            template_df.to_excel(writer, sheet_name=update.sheet_name, index=False)

        return {"message": f"'{update.sheet_name}' was reset to template version"}

    except Exception as e:
        return {"error": f"Failed to reset: {str(e)}"}
    

# Techno-Economic Inputs
@app.get("/technoeconomic-inputs/{model}")
async def get_technoeconomic_inputs(model: str):
    if model.lower() in ["none", "bau"]:
        if not os.path.exists(TECHNOECONOMIC_INPUTS_TEMPLATE):
            return {"error": "Template file is missing"}
        
        return FileResponse(
            TECHNOECONOMIC_INPUTS_TEMPLATE,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=TECHNOECONOMIC_INPUTS_TEMPLATE
        )

    FULL_TECHNO_MODEL_PATH = f"{TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx"

    if not os.path.exists(TECHNOECONOMIC_INPUTS_TEMPLATE):
        return {"error": "Template file is missing"}

    if not os.path.exists(FULL_TECHNO_MODEL_PATH):
        shutil.copy(TECHNOECONOMIC_INPUTS_TEMPLATE, FULL_TECHNO_MODEL_PATH)

    return FileResponse(
        FULL_TECHNO_MODEL_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=FULL_TECHNO_MODEL_PATH
    )

@app.post("/save-technoeconomic-inputs")
async def save_technoeconomic_inputs(update: SheetUpdate):

    if not update.model or not update.sheet_name or update.data is None:
        return {"error": "Missing model, sheet_name or data"}

    FULL_TECHNO_MODEL_PATH = f"{TECHNOECONOMIC_INPUTS_MODEL}{update.model}.xlsx"
    df_new = pd.DataFrame(update.data)

    with pd.ExcelWriter(FULL_TECHNO_MODEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_new.to_excel(writer, sheet_name=update.sheet_name, index=False)

    return {"message": f"Sheet '{update.sheet_name}' saved successfully for model '{update.model}'"}

@app.post("/reset-technoeconomic-inputs")
async def reset_technoeconomic_inputs(update: SheetUpdate):
    model = update.model
    sheet_name = update.sheet_name

    if not model or not sheet_name:
        raise HTTPException(status_code=400, detail="Missing model or sheet_name")

    FULL_TECHNO_MODEL_PATH = f"{TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx"

    if not os.path.exists(FULL_TECHNO_MODEL_PATH):
        if not os.path.exists(TECHNOECONOMIC_INPUTS_TEMPLATE):
            raise HTTPException(status_code=500, detail="Template file is missing")
        shutil.copy(TECHNOECONOMIC_INPUTS_TEMPLATE, FULL_TECHNO_MODEL_PATH)

    try:
        with pd.ExcelFile(TECHNOECONOMIC_INPUTS_TEMPLATE, engine='openpyxl') as xls:
            if sheet_name not in xls.sheet_names:
                raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found in template")
            df_template = pd.read_excel(xls, sheet_name=sheet_name, engine='openpyxl')

        with pd.ExcelWriter(FULL_TECHNO_MODEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_template.to_excel(writer, sheet_name=sheet_name, index=False)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error resetting sheet: {e}")

    return {"message": f"Sheet '{sheet_name}' reset to template version for model '{model}'"}


# Carbon Credits
@app.get("/carbon-credits")
def get_carbon_credits():

    if not os.path.exists(CARBON_CREDITS_PATH):
        return {"error": "'Carbon credits' file is missing, BAU model was not initialized properly."}

    wb = load_workbook(CARBON_CREDITS_PATH)
    ws = wb["Carbon Credits"]

    model_names = set()
    pattern = r"CO2 emited\s*-\s*(.+?)\s*scenario"

    for _, row in enumerate(ws.iter_rows(), start=1):
        cell = row[0].value if len(row) > 0 else None
        if isinstance(cell, str):
            cell_cleaned = re.sub(r"\{.*?\}", "", cell)
            match = re.search(pattern, cell_cleaned)
            if match:
                model = match.group(1).strip()
                if model:
                    model_names.add(model)
    
    models_sorted = sorted(model_names)
    if "BAU" in models_sorted:
        models_sorted.remove("BAU")

    e.apply_formulas(CARBON_CREDITS_PATH, FORMULAS_JSON_PATH, models_sorted, [], [])

    return FileResponse(
        CARBON_CREDITS_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="carbon-credits.xlsx"
    )

@app.post("/save-carbon-credits")
async def save_carbon_credits(update: SheetUpdate):

    df_new = pd.DataFrame(update.data)

    if 'index' in df_new.columns:
        df_new.set_index('index', inplace=True)

    excel_path = CARBON_CREDITS_PATH
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_new.to_excel(writer, sheet_name=update.sheet_name, index=False)

    return {"message": "The file with the fuel market information was updated successfully"}

@app.post("/reset-carbon-credits")
def reset_carbon_credits(update: SheetUpdate):
    
    if not os.path.exists(CARBON_CREDITS_PATH):
        return {"error": "Fuel market information file does not exist"}
    
    if not os.path.exists(CARBON_CREDITS_TEMPLATE_PATH):
        return {"error": "Template file is missing"}
    
    try:
        template_df = pd.read_excel(CARBON_CREDITS_TEMPLATE_PATH, sheet_name=update.sheet_name)
        wb = load_workbook(CARBON_CREDITS_PATH)

        if update.sheet_name in wb.sheetnames:
            std = wb[update.sheet_name]
            wb.remove(std)

        wb.save(CARBON_CREDITS_PATH)

        with pd.ExcelWriter(CARBON_CREDITS_PATH, engine='openpyxl', mode='a') as writer:
            template_df.to_excel(writer, sheet_name=update.sheet_name, index=False)

        return {"message": f"'{update.sheet_name}' was reset to template version"}

    except Exception as e:
        return {"error": f"Failed to reset: {str(e)}"}


# Financial Statements
@app.get("/financial-statements/{model}")
async def get_financial_statements(model: str):
    global START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS
    if model.lower() in ["none", "bau"]:
        if not os.path.exists(FINANCIAL_STATEMENTS_TEMPLATE):
            return {"error": "Template file is missing"}
        
        return FileResponse(
            FINANCIAL_STATEMENTS_TEMPLATE,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=FINANCIAL_STATEMENTS_TEMPLATE
        )

    FULL_FFSS_MODEL_PATH = f"{FINANCIAL_STATEMENTS_MODEL}{model}.xlsx"

    if not os.path.exists(FINANCIAL_STATEMENTS_TEMPLATE):
        return {"error": "Template file is missing"}

    if not os.path.exists(FULL_FFSS_MODEL_PATH):
        shutil.copy(FINANCIAL_STATEMENTS_TEMPLATE, FULL_FFSS_MODEL_PATH)

    if START_YEAR_TECHNO_MODELS is None and END_YEAR_TECHNO_MODELS is None:
        path = f"{TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx"
        START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS = detect_year_range(path)

    xls_fuel_market = pd.ExcelFile(FUEL_MARKET_INFORMATION_PATH, engine="openpyxl")
    fuel_market_sheets = xls_fuel_market.sheet_names
    expected_sheets = [
        "Electricity (Low access)" if name.strip().lower() == "carbon"
        else "Electricity" if name.strip().lower() == "electricity"
        else name
        for name in fuel_market_sheets
    ]
    expected_sheets.append("E-Cooking")
    priority_order = ["E-Cooking", "Electricity", "Electricity (Low access)"]
    expected_sheets = sorted(expected_sheets, key=lambda x: priority_order.index(x) if x in priority_order else len(priority_order))

    xls_model = pd.ExcelFile(FULL_FFSS_MODEL_PATH, engine="openpyxl")
    existing_sheets = xls_model.sheet_names

    xls_template = pd.ExcelFile(FINANCIAL_STATEMENTS_TEMPLATE, engine="openpyxl")
    df_ecooking = pd.read_excel(xls_template, sheet_name="E-Cooking", engine="openpyxl", header=None)

    missing_sheets = [s for s in expected_sheets if s not in existing_sheets]

    if missing_sheets:
        sheets_dict = {
            sheet: pd.read_excel(xls_model, sheet_name=sheet)
            for sheet in existing_sheets
        }
        for missing in missing_sheets:
            sheets_dict[missing] = df_ecooking.copy()
        with pd.ExcelWriter(FULL_FFSS_MODEL_PATH, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(FULL_FFSS_MODEL_PATH)

    sheets_to_remove = [ws.title for ws in wb.worksheets if ws.title not in expected_sheets]

    for sheet_name in sheets_to_remove:
        std = wb[sheet_name]
        wb.remove(std)

    for ws in wb.worksheets:
        year_row = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, str) and cell_val.strip().lower() == "baseline":
                    year_row = row
                    break

            if year_row:
                break
        
        if year_row is None:
            continue

        cols_to_delete = []
        for col in range(3, ws.max_column + 1):
            cell_val = ws.cell(row=year_row, column=col).value
            try:
                year = int(cell_val)
                if year <= START_YEAR_TECHNO_MODELS or year > END_YEAR_TECHNO_MODELS:
                    cols_to_delete.append(col)
            except (TypeError, ValueError):
                pass

        for col_idx in reversed(cols_to_delete):
            ws.delete_cols(col_idx)

    wb.save(FULL_FFSS_MODEL_PATH)

    if not os.path.exists(CARBON_CREDITS_PATH):
        return {"error": "'Carbon credits' file is missing, BAU model was not initialized properly."}

    wb = load_workbook(CARBON_CREDITS_PATH)
    ws = wb["Carbon Credits"]

    model_names = set()
    pattern = r"CO2 emited\s*-\s*(.+?)\s*scenario"

    for _, row in enumerate(ws.iter_rows(), start=1):
        cell = row[0].value if len(row) > 0 else None
        if isinstance(cell, str):
            cell_cleaned = re.sub(r"\{.*?\}", "", cell)
            match = re.search(pattern, cell_cleaned)
            if match:
                model = match.group(1).strip()
                if model:
                    model_names.add(model)
    
    models_sorted = sorted(model_names)
    if "BAU" in models_sorted:
        models_sorted.remove("BAU")
    
    e.apply_formulas(FULL_FFSS_MODEL_PATH, FORMULAS_JSON_PATH, models_sorted, fuel_market_sheets, expected_sheets)

    return FileResponse(
        FULL_FFSS_MODEL_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=FULL_FFSS_MODEL_PATH
    )


# CAPEX Fuel Markets
@app.get("/capex-fuel-market/{model}")
async def get_capex_market(model: str):
    global START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS
    if model.lower() in ["none", "bau"]:
        if not os.path.exists(CAPEX_MARKET_TEMPLATE):
            return {"error": "Template file is missing"}
        
        return FileResponse(
            CAPEX_MARKET_TEMPLATE,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=CAPEX_MARKET_TEMPLATE
        )

    FULL_CAPEX_MODEL_PATH = f"{CAPEX_MARKET_MODEL}{model}.xlsx"

    if not os.path.exists(CAPEX_MARKET_TEMPLATE):
        return {"error": "Template file is missing"}

    if not os.path.exists(FULL_CAPEX_MODEL_PATH):
        shutil.copy(CAPEX_MARKET_TEMPLATE, FULL_CAPEX_MODEL_PATH)

    if START_YEAR_TECHNO_MODELS is None and END_YEAR_TECHNO_MODELS is None:
        path = f"{TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx"
        START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS = detect_year_range(path)

    xls_fuel_market = pd.ExcelFile(FUEL_MARKET_INFORMATION_PATH, engine="openpyxl")
    fuel_market_sheets = xls_fuel_market.sheet_names
    expected_sheets = [name for name in fuel_market_sheets if name.strip().lower() != "carbon" ]

    xls_model = pd.ExcelFile(FULL_CAPEX_MODEL_PATH, engine="openpyxl")
    existing_sheets = xls_model.sheet_names

    xls_template = pd.ExcelFile(CAPEX_MARKET_TEMPLATE, engine="openpyxl")
    df_lpg = pd.read_excel(xls_template, sheet_name="LPG", engine="openpyxl", header=None)

    missing_sheets = [s for s in expected_sheets if s not in existing_sheets]

    if missing_sheets:
        sheets_dict = {
            sheet: pd.read_excel(xls_model, sheet_name=sheet)
            for sheet in existing_sheets
        }
        for missing in missing_sheets:
            sheets_dict[missing] = df_lpg.copy()
        with pd.ExcelWriter(FULL_CAPEX_MODEL_PATH, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(FULL_CAPEX_MODEL_PATH)

    sheets_to_remove = [ws.title for ws in wb.worksheets if ws.title not in expected_sheets]

    for sheet_name in sheets_to_remove:
        std = wb[sheet_name]
        wb.remove(std)

    for ws in wb.worksheets:
        year_row = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, str) and cell_val.strip().lower() == "baseline":
                    year_row = row
                    break

            if year_row:
                break
        
        if year_row is None:
            continue

        cols_to_delete = []
        for col in range(3, ws.max_column + 1):
            cell_val = ws.cell(row=year_row, column=col).value
            try:
                year = int(cell_val)
                if year <= START_YEAR_TECHNO_MODELS or year > END_YEAR_TECHNO_MODELS:
                    cols_to_delete.append(col)
            except (TypeError, ValueError):
                pass

        for col_idx in reversed(cols_to_delete):
            ws.delete_cols(col_idx)

    wb.save(FULL_CAPEX_MODEL_PATH)

    return FileResponse(
        FULL_CAPEX_MODEL_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=FULL_CAPEX_MODEL_PATH
    )


