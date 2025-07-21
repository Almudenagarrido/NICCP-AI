from fastapi import APIRouter
from fastapi.responses import FileResponse
from openpyxl import load_workbook
import pandas as pd
import re

from models import SheetUpdate
import excel_utils as e
import config as c

router = APIRouter()

@router.get("/carbon-credits")
def get_carbon_credits():

    if not pd.os.path.exists(c.CARBON_CREDITS_PATH):
        return {"error": "'Carbon credits' file is missing, BAU model was not initialized properly."}

    wb = load_workbook(c.CARBON_CREDITS_PATH)
    ws = wb["Carbon Credits"]

    model_names = set()
    pattern = r"CO2 emited\s*-\s*(.+?)\s*scenario"

    for _, row in enumerate(ws.iter_rows(), start=1):
        cell = row[0].value if len(row) > 0 else None
        if isinstance(cell, str):
            cell_cleaned = re.sub(r"\{.*?\}", "", cell)
            match = re.search(pattern, cell_cleaned)
            if match:
                model = match.group(1).strip()
                if model:
                    model_names.add(model)

    models_sorted = sorted(model_names)
    if "BAU" in models_sorted:
        models_sorted.remove("BAU")

    e.apply_formulas(c.CARBON_CREDITS_PATH, c.FORMULAS_JSON_PATH, models_sorted, [], [])

    return FileResponse(
        c.CARBON_CREDITS_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="carbon-credits.xlsx"
    )

@router.post("/save-carbon-credits")
async def save_carbon_credits(update: SheetUpdate):

    df_new = pd.DataFrame(update.data)

    if 'index' in df_new.columns:
        df_new.set_index('index', inplace=True)

    excel_path = c.CARBON_CREDITS_PATH
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_new.to_excel(writer, sheet_name=update.sheet_name, index=False)

    return {"message": "The file with the fuel market information was updated successfully"}

@router.post("/reset-carbon-credits")
def reset_carbon_credits(update: SheetUpdate):

    if not pd.os.path.exists(c.CARBON_CREDITS_PATH):
        return {"error": "Fuel market information file does not exist"}

    if not pd.os.path.exists(c.CARBON_CREDITS_TEMPLATE_PATH):
        return {"error": "Template file is missing"}

    try:
        template_df = pd.read_excel(c.CARBON_CREDITS_TEMPLATE_PATH, sheet_name=update.sheet_name)
        wb = load_workbook(c.CARBON_CREDITS_PATH)

        if update.sheet_name in wb.sheetnames:
            std = wb[update.sheet_name]
            wb.remove(std)

        wb.save(c.CARBON_CREDITS_PATH)

        with pd.ExcelWriter(c.CARBON_CREDITS_PATH, engine='openpyxl', mode='a') as writer:
            template_df.to_excel(writer, sheet_name=update.sheet_name, index=False)

        return {"message": f"'{update.sheet_name}' was reset to template version"}

    except Exception as e:
        return {"error": f"Failed to reset: {str(e)}"}
