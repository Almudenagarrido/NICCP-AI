from fastapi import APIRouter
from fastapi.responses import FileResponse
from openpyxl import load_workbook
import pandas as pd
import shutil
import os

from models import SheetUpdate
from utils import detect_year_range
import config as c

router = APIRouter()

@router.get("/design-capital-structure/{model}")
async def get_design_capital_structure(model: str):
    if model.lower() in ["none", "bau"]:
        if not os.path.exists(c.DESIGN_CAPITAL_STRUCTURE_TEMPLATE):
            return {"error": "Template file is missing"}

        return FileResponse(
            c.DESIGN_CAPITAL_STRUCTURE_TEMPLATE,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=c.DESIGN_CAPITAL_STRUCTURE_TEMPLATE
        )

    FULL_DESIGN_MODEL_PATH = f"{c.DESIGN_CAPITAL_STRUCTURE_MODEL}{model}.xlsx"

    if not os.path.exists(c.FUEL_MARKET_INFORMATION_PATH):
        return {"error": "fuel-market-information.xlsx not found"}

    if not os.path.exists(c.DESIGN_CAPITAL_STRUCTURE_TEMPLATE):
        return {"error": "Template file is missing"}

    if not os.path.exists(FULL_DESIGN_MODEL_PATH):
        shutil.copy(c.DESIGN_CAPITAL_STRUCTURE_TEMPLATE, FULL_DESIGN_MODEL_PATH)

    if c.START_YEAR_TECHNO_MODELS is None and c.END_YEAR_TECHNO_MODELS is None:
        path = f"{c.TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx"
        c.START_YEAR_TECHNO_MODELS, c.END_YEAR_TECHNO_MODELS = detect_year_range(path)

    xls_fuel_market = pd.ExcelFile(c.FUEL_MARKET_INFORMATION_PATH, engine="openpyxl")
    fuel_market_sheets = xls_fuel_market.sheet_names
    expected_sheets = [
        "Electricity (Low access)" if name.strip().lower() == "carbon"
        else "Electricity & E-Cooking" if name.strip().lower() == "electricity"
        else name
        for name in fuel_market_sheets
    ]

    xls_model = pd.ExcelFile(FULL_DESIGN_MODEL_PATH, engine="openpyxl")
    existing_sheets = xls_model.sheet_names

    xls_template = pd.ExcelFile(c.DESIGN_CAPITAL_STRUCTURE_TEMPLATE, engine="openpyxl")
    df_ecooking = pd.read_excel(xls_template, sheet_name="Electricity & E-Cooking", engine="openpyxl", header=None)

    missing_sheets = [s for s in expected_sheets if s not in existing_sheets]

    if missing_sheets:
        sheets_dict = {
            sheet: pd.read_excel(xls_model, sheet_name=sheet)
            for sheet in existing_sheets
        }
        for missing in missing_sheets:
            sheets_dict[missing] = df_ecooking.copy()
        with pd.ExcelWriter(FULL_DESIGN_MODEL_PATH, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(FULL_DESIGN_MODEL_PATH)

    sheets_to_remove = [ws.title for ws in wb.worksheets if ws.title not in expected_sheets]

    for sheet_name in sheets_to_remove:
        std = wb[sheet_name]
        wb.remove(std)

    for ws in wb.worksheets:
        year_row = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, str) and cell_val.strip().lower() == "baseline":
                    year_row = row
                    break

            if year_row:
                break

        if year_row is None:
            continue

        cols_to_delete = []
        for col in range(3, ws.max_column + 1):
            cell_val = ws.cell(row=year_row, column=col).value
            try:
                year = int(cell_val)
                if year <= c.START_YEAR_TECHNO_MODELS or year > c.END_YEAR_TECHNO_MODELS:
                    cols_to_delete.append(col)
            except (TypeError, ValueError):
                pass

        for col_idx in reversed(cols_to_delete):
            ws.delete_cols(col_idx)

    wb.save(FULL_DESIGN_MODEL_PATH)

    return FileResponse(
        FULL_DESIGN_MODEL_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=FULL_DESIGN_MODEL_PATH
    )

@router.post("/save-design-capital-structure")
async def save_design_capital_structure(update: SheetUpdate):
    FULL_DESIGN_MODEL_PATH = f"{c.DESIGN_CAPITAL_STRUCTURE_MODEL}{update.model}.xlsx"

    if not os.path.exists(FULL_DESIGN_MODEL_PATH):
        return {"error": f"File for model '{update.model}' not found"}

    try:
        df = pd.DataFrame(update.data)
        if 'index' in df.columns:
            df.set_index('index', inplace=True)

        with pd.ExcelWriter(FULL_DESIGN_MODEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=update.sheet_name, index=False)

        return {"message": f"Design capital structure for model '{update.model}' saved successfully."}
    except Exception as e:
        return {"error": f"Failed to save: {str(e)}"}

@router.post("/reset-design-capital-structure")
def reset_design_capital_structure(update: SheetUpdate):
    FULL_DESIGN_MODEL_PATH = f"{c.DESIGN_CAPITAL_STRUCTURE_MODEL}{update.model}.xlsx"

    if not os.path.exists(FULL_DESIGN_MODEL_PATH):
        return {"error": f"Model file '{FULL_DESIGN_MODEL_PATH}' not found"}

    if not os.path.exists(c.DESIGN_CAPITAL_STRUCTURE_TEMPLATE):
        return {"error": f"Template file '{c.DESIGN_CAPITAL_STRUCTURE_TEMPLATE}' not found"}

    try:
        if update.sheet_name not in {"Electricity", "LPG"}:
            template_sheet = "Electricity & E-Cooking"
        else:
            template_sheet = update.sheet_name

        template_df = pd.read_excel(c.DESIGN_CAPITAL_STRUCTURE_TEMPLATE, sheet_name=template_sheet)
        wb = load_workbook(FULL_DESIGN_MODEL_PATH)

        if update.sheet_name in wb.sheetnames:
            std = wb[update.sheet_name]
            wb.remove(std)
            wb.save(FULL_DESIGN_MODEL_PATH)

        with pd.ExcelWriter(FULL_DESIGN_MODEL_PATH, engine='openpyxl', mode='a') as writer:
            template_df.to_excel(writer, sheet_name=update.sheet_name, index=False)

        return {"message": f"'{update.sheet_name}' was reset to template version"}

    except Exception as e:
        return {"error": f"Failed to reset: {str(e)}"}
