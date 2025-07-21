from fastapi import APIRouter
from fastapi.responses import FileResponse
from openpyxl import load_workbook
import pandas as pd
import shutil
import os
import re

from utils import detect_year_range
import excel_utils as e
import config as c

router = APIRouter()

@router.get("/financial-statements/{model}")
async def get_financial_statements(model: str):
    if model.lower() in ["none", "bau"]:
        if not os.path.exists(c.FINANCIAL_STATEMENTS_TEMPLATE):
            return {"error": "Template file is missing"}

        return FileResponse(
            c.FINANCIAL_STATEMENTS_TEMPLATE,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=c.FINANCIAL_STATEMENTS_TEMPLATE
        )

    FULL_FFSS_MODEL_PATH = f"{c.FINANCIAL_STATEMENTS_MODEL}{model}.xlsx"

    if not os.path.exists(c.FINANCIAL_STATEMENTS_TEMPLATE):
        return {"error": "Template file is missing"}

    if not os.path.exists(FULL_FFSS_MODEL_PATH):
        shutil.copy(c.FINANCIAL_STATEMENTS_TEMPLATE, FULL_FFSS_MODEL_PATH)

    if c.START_YEAR_TECHNO_MODELS is None and c.END_YEAR_TECHNO_MODELS is None:
        path = f"{c.TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx"
        c.START_YEAR_TECHNO_MODELS, c.END_YEAR_TECHNO_MODELS = detect_year_range(path)

    xls_fuel_market = pd.ExcelFile(c.FUEL_MARKET_INFORMATION_PATH, engine="openpyxl")
    fuel_market_sheets = xls_fuel_market.sheet_names
    expected_sheets = [
        "Electricity (Low access)" if name.strip().lower() == "carbon"
        else "Electricity" if name.strip().lower() == "electricity"
        else name
        for name in fuel_market_sheets
    ]
    expected_sheets.append("E-Cooking")
    priority_order = ["E-Cooking", "Electricity", "Electricity (Low access)"]
    expected_sheets = sorted(expected_sheets, key=lambda x: priority_order.index(x) if x in priority_order else len(priority_order))

    xls_model = pd.ExcelFile(FULL_FFSS_MODEL_PATH, engine="openpyxl")
    existing_sheets = xls_model.sheet_names

    xls_template = pd.ExcelFile(c.FINANCIAL_STATEMENTS_TEMPLATE, engine="openpyxl")
    df_ecooking = pd.read_excel(xls_template, sheet_name="E-Cooking", engine="openpyxl", header=None)

    missing_sheets = [s for s in expected_sheets if s not in existing_sheets]

    if missing_sheets:
        sheets_dict = {
            sheet: pd.read_excel(xls_model, sheet_name=sheet)
            for sheet in existing_sheets
        }
        for missing in missing_sheets:
            sheets_dict[missing] = df_ecooking.copy()
        with pd.ExcelWriter(FULL_FFSS_MODEL_PATH, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(FULL_FFSS_MODEL_PATH)

    sheets_to_remove = [ws.title for ws in wb.worksheets if ws.title not in expected_sheets]

    for sheet_name in sheets_to_remove:
        std = wb[sheet_name]
        wb.remove(std)

    for ws in wb.worksheets:
        year_row = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, str) and cell_val.strip().lower() == "baseline":
                    year_row = row
                    break

            if year_row:
                break

        if year_row is None:
            continue

        cols_to_delete = []
        for col in range(3, ws.max_column + 1):
            cell_val = ws.cell(row=year_row, column=col).value
            try:
                year = int(cell_val)
                if year <= c.START_YEAR_TECHNO_MODELS or year > c.END_YEAR_TECHNO_MODELS:
                    cols_to_delete.append(col)
            except (TypeError, ValueError):
                pass

        for col_idx in reversed(cols_to_delete):
            ws.delete_cols(col_idx)

    wb.save(FULL_FFSS_MODEL_PATH)

    if not os.path.exists(c.CARBON_CREDITS_PATH):
        return {"error": "'Carbon credits' file is missing, BAU model was not initialized properly."}

    wb = load_workbook(c.CARBON_CREDITS_PATH)
    ws = wb["Carbon Credits"]

    model_names = set()
    pattern = r"CO2 emited\s*-\s*(.+?)\s*scenario"

    for _, row in enumerate(ws.iter_rows(), start=1):
        cell = row[0].value if len(row) > 0 else None
        if isinstance(cell, str):
            cell_cleaned = re.sub(r"\{.*?\}", "", cell)
            match = re.search(pattern, cell_cleaned)
            if match:
                model = match.group(1).strip()
                if model:
                    model_names.add(model)

    models_sorted = sorted(model_names)
    if "BAU" in models_sorted:
        models_sorted.remove("BAU")

    e.apply_formulas(FULL_FFSS_MODEL_PATH, c.FORMULAS_JSON_PATH, models_sorted, fuel_market_sheets, expected_sheets)

    return FileResponse(
        FULL_FFSS_MODEL_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=FULL_FFSS_MODEL_PATH
    )
