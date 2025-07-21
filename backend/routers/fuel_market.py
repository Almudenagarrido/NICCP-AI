from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from openpyxl import load_workbook
import pandas as pd
import shutil
import os

from models import SheetUpdate, MarketName
import config as c

router = APIRouter()

@router.get("/fuel-market-information")
def get_fuel_market_information():
    global START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS

    models = [
        os.path.splitext(f)[0].removeprefix("technoeconomic-inputs-")
        for f in os.listdir(c.TECHNOECONOMIC_INPUTS_FOLDER)
        if f.lower().endswith(tuple(ext.lower() for ext in c.VALID_EXTENSIONS)) and "template" not in f.lower()
    ]

    if not os.path.exists(c.FUEL_MARKET_INFORMATION_PATH):
        if os.path.exists(c.FUEL_MARKET_INFORMATION_TEMPLATE):
            shutil.copy(c.FUEL_MARKET_INFORMATION_TEMPLATE, c.FUEL_MARKET_INFORMATION_PATH)
        else:
            return {"error": "Template file is missing"}

    wb = load_workbook(c.FUEL_MARKET_INFORMATION_PATH)

    for ws in wb.worksheets:
        year_row = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, str) and cell_val.strip().lower() == "baseline":
                    year_row = row
                    break
            if year_row:
                break

        if year_row is None:
            continue

        cols_to_delete = []
        for col in range(3, ws.max_column + 1):
            cell_val = ws.cell(row=year_row, column=col).value
            try:
                year = int(cell_val)
                if year <= c.START_YEAR_TECHNO_MODELS or year > c.END_YEAR_TECHNO_MODELS:
                    cols_to_delete.append(col)
            except (TypeError, ValueError):
                pass

        for col_idx in reversed(cols_to_delete):
            ws.delete_cols(col_idx)

        if ws.title == "LPG":
            existing_rows = {}

            for i in range(2, ws.max_row + 1):
                val = ws.cell(row=i, column=1).value
                if val:
                    key = str(val).strip()
                    existing_rows[key] = i

            expected_rows = {
                f"% EBITDA Margin - {model}" for model in models
            }.union({
                f"OPEX subsidies - {model}" for model in models
            })

            for row_label, row_num in list(existing_rows.items())[::-1]:
                if (row_label.startswith("% EBITDA Margin - ") or row_label.startswith("OPEX subsidies - ")) and row_label not in expected_rows:
                    ws.delete_rows(row_num)

            existing_inputs = {
                str(ws.cell(row=i, column=1).value).strip()
                for i in range(2, ws.max_row + 1)
                if ws.cell(row=i, column=1).value
            }

            for model_name in models:
                for row_label in [f"% EBITDA Margin - {model_name}", f"OPEX subsidies - {model_name}"]:
                    if row_label not in existing_inputs:
                        new_row = ws.max_row + 1
                        ws.cell(row=new_row, column=1).value = row_label
                        ws.cell(row=new_row, column=2).value = "%"
                        for col in range(3, ws.max_column + 1):
                            ws.cell(row=new_row, column=col).value = 0

    wb.save(c.FUEL_MARKET_INFORMATION_PATH)

    return FileResponse(
        c.FUEL_MARKET_INFORMATION_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="fuel-market-information.xlsx"
    )

@router.post("/save-fuel-market-information")
async def save_fuel_market_information(update: SheetUpdate):

    df_new = pd.DataFrame(update.data)

    if 'index' in df_new.columns:
        df_new.set_index('index', inplace=True)

    excel_path = c.FUEL_MARKET_INFORMATION_PATH
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_new.to_excel(writer, sheet_name=update.sheet_name, index=False)

    return {"message": "The file with the fuel market information was updated successfully"}

@router.post("/reset-fuel-market-information")
def reset_fuel_market_information(update: SheetUpdate):

    if not os.path.exists(c.FUEL_MARKET_INFORMATION_PATH):
        return {"error": "Fuel market information file does not exist"}

    if not os.path.exists(c.FUEL_MARKET_INFORMATION_TEMPLATE):
        return {"error": "Template file is missing"}

    try:
        if update.sheet_name not in {"Carbon", "LPG"}:
            template_sheet = "Electricity"
        else:
            template_sheet = update.sheet_name

        template_df = pd.read_excel(c.FUEL_MARKET_INFORMATION_TEMPLATE, sheet_name=template_sheet)
        wb = load_workbook(c.FUEL_MARKET_INFORMATION_PATH)

        if update.sheet_name in wb.sheetnames:
            std = wb[update.sheet_name]
            wb.remove(std)

        wb.save(c.FUEL_MARKET_INFORMATION_PATH)

        with pd.ExcelWriter(c.FUEL_MARKET_INFORMATION_PATH, engine='openpyxl', mode='a') as writer:
            template_df.to_excel(writer, sheet_name=update.sheet_name, index=False)

        return {"message": f"'{update.sheet_name}' was reset to template version"}

    except Exception as e:
        return {"error": f"Failed to reset: {str(e)}"}

@router.post("/add-fuel-market")
def add_fuel_market(market: MarketName):
    if not os.path.exists(c.FUEL_MARKET_INFORMATION_PATH):
        return {"error": "fuel-market-information.xlsx not found"}

    xls_template = pd.ExcelFile(c.FUEL_MARKET_INFORMATION_TEMPLATE, engine="openpyxl")
    xls = pd.ExcelFile(c.FUEL_MARKET_INFORMATION_PATH, engine="openpyxl")
    df_ecooking = pd.read_excel(xls_template, sheet_name="Electricity", engine="openpyxl")
    new_sheet_name = market.name

    if "Electricity" not in xls_template.sheet_names:
        return {"error": "Base sheet 'Electricity' not found"}

    if new_sheet_name in xls.sheet_names:
        return {"error": f"The sheet '{new_sheet_name}' already exists."}

    sheets_dict = {sheet: xls.parse(sheet) for sheet in xls.sheet_names}
    sheets_dict[new_sheet_name] = df_ecooking.copy()

    with pd.ExcelWriter(c.FUEL_MARKET_INFORMATION_PATH, engine="openpyxl", mode="w") as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    return {"message": f"Market '{new_sheet_name}' added successfully."}

@router.post("/delete-fuel-market")
def delete_fuel_market(market: MarketName):
    if not os.path.exists(c.FUEL_MARKET_INFORMATION_PATH):
        return {"error": "fuel-market-information.xlsx not found"}

    xls = pd.ExcelFile(c.FUEL_MARKET_INFORMATION_PATH, engine="openpyxl")
    sheet_to_delete = market.name

    if sheet_to_delete not in xls.sheet_names:
        return {"error": f"The sheet '{sheet_to_delete}' does not exist."}

    remaining_sheets = {
        sheet: xls.parse(sheet)
        for sheet in xls.sheet_names if sheet != sheet_to_delete
    }

    with pd.ExcelWriter(c.FUEL_MARKET_INFORMATION_PATH, engine="openpyxl", mode="w") as writer:
        for name, df in remaining_sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

    return {"message": f"Market '{sheet_to_delete}' deleted successfully."}
