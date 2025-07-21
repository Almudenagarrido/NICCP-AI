from fastapi import APIRouter, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.responses import FileResponse
from openpyxl import load_workbook
import pandas as pd
import shutil
import os
import re
import tempfile
from io import BytesIO

import excel_utils as e
from utils import detect_year_range
import config as c

router = APIRouter()

@router.get("/technoeconomic-models")
def list_techno_models():
    if not os.path.exists(c.CARBON_CREDITS_PATH):
        return {"models": []}

    wb = load_workbook(c.CARBON_CREDITS_PATH)
    ws = wb["Carbon Credits"]

    model_names = set()
    pattern = r"CO2 emited\s*-\s*(.+?)\s*scenario"

    for _, row in enumerate(ws.iter_rows(), start=1):
        cell = row[0].value if len(row) > 0 else None
        if isinstance(cell, str):
            cell_cleaned = re.sub(r"\{.*?\}", "", cell)
            match = re.search(pattern, cell_cleaned)
            if match:
                model = match.group(1).strip()
                if model:
                    model_names.add(model)

    models_sorted = sorted(model_names)
    if "BAU" in models_sorted:
        models_sorted.remove("BAU")
        models_sorted.insert(0, "BAU")

    return {"models": models_sorted}

@router.post("/create-technoeconomic-model/{model}")
async def create_technoeconomic_model(model: str, start_year: int, end_year: int):
    global START_YEAR_TECHNO_MODELS, END_YEAR_TECHNO_MODELS

    model = model.strip()
    if not model:
        return {"error": "Model name cannot be empty"}

    if start_year >= end_year:
        return {"error": "Start year must be less than end year"}

    if model.lower() == "bau":
        file_path = c.CARBON_CREDITS_PATH
        template_path = c.CARBON_CREDITS_TEMPLATE_PATH
    else:
        file_path = f"{c.TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx"
        template_path = c.TECHNOECONOMIC_INPUTS_TEMPLATE

    if not os.path.exists(file_path):
        if not os.path.exists(template_path):
            return {"error": f"Template file missing: {template_path}"}
        shutil.copy(template_path, file_path)

    if model != "bau":
        wb_cc = load_workbook(c.CARBON_CREDITS_PATH)
        ws_cc = wb_cc["Carbon Credits"]

        rows_to_insert = []
        for row_idx in range(1, ws_cc.max_row + 1):
            cell_val = ws_cc.cell(row=row_idx, column=1).value
            if cell_val and isinstance(cell_val, str) and "{model}" in cell_val:
                row_values = [ws_cc.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws_cc.max_column + 1)]
                rows_to_insert.append((row_idx, row_values))

        for orig_row_idx, row_values in reversed(rows_to_insert):
            insert_at = orig_row_idx + 1
            ws_cc.insert_rows(insert_at)

            for col_idx, val in enumerate(row_values, start=1):
                if col_idx == 1 and isinstance(val, str):
                    val = val.replace("{model}", model)
                ws_cc.cell(row=insert_at, column=col_idx, value=val)

        wb_cc.save(c.CARBON_CREDITS_PATH)

        ref_start, ref_end = detect_year_range(c.CARBON_CREDITS_PATH)
        if ref_start != start_year or ref_end != end_year:
            return {
                "error": f"Year range mismatch: expected {ref_start}-{ref_end}, got {start_year}-{end_year}"
            }

    if c.START_YEAR_TECHNO_MODELS is None or c.END_YEAR_TECHNO_MODELS is None:
        c.START_YEAR_TECHNO_MODELS = start_year
        c.END_YEAR_TECHNO_MODELS = end_year

    wb = load_workbook(file_path)

    if model == "bau":
        ws = wb["Carbon Credits"]
        year_row = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, str) and cell_val.strip().lower() == "baseline":
                    year_row = row
                    break
            if year_row:
                break

        if year_row:
            cols_to_delete = []
            for col in range(3, ws.max_column + 1):
                cell_val = ws.cell(row=year_row, column=col).value
                try:
                    year = int(cell_val)
                    if year <= c.START_YEAR_TECHNO_MODELS or year > c.END_YEAR_TECHNO_MODELS:
                        cols_to_delete.append(col)
                except (TypeError, ValueError):
                    continue

            for col_idx in reversed(cols_to_delete):
                ws.delete_cols(col_idx)
    else:
        xls_fuel_market = pd.ExcelFile(c.FUEL_MARKET_INFORMATION_PATH, engine="openpyxl")
        fuel_market_sheets = xls_fuel_market.sheet_names
        expected_sheets = [
            "C02" if name.strip().lower() == "carbon" else name
            for name in fuel_market_sheets
        ]

        xls_model = pd.ExcelFile(file_path, engine="openpyxl")
        existing_sheets = xls_model.sheet_names

        xls_template = pd.ExcelFile(template_path, engine="openpyxl")
        df_template = pd.read_excel(xls_template, sheet_name="LPG", engine="openpyxl")

        missing_sheets = [s for s in expected_sheets if s not in existing_sheets]

        if missing_sheets:
            sheets_dict = {
                sheet: pd.read_excel(xls_model, sheet_name=sheet)
                for sheet in existing_sheets
            }
            for missing in missing_sheets:
                sheets_dict[missing] = df_template.copy()
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                for sheet_name, df in sheets_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        sheets_to_remove = [
            ws.title for ws in wb.worksheets
            if ws.title not in expected_sheets and ws.title != "Rest of subsidies or taxes"
        ]
        for sheet_name in sheets_to_remove:
            wb.remove(wb[sheet_name])

        for ws in wb.worksheets:
            year_row = None
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=row, column=col).value
                    if isinstance(cell_val, str) and cell_val.strip().lower() == "baseline":
                        year_row = row
                        break
                if year_row:
                    break

            if year_row is None:
                continue

            cols_to_delete = []
            for col in range(3, ws.max_column + 1):
                cell_val = ws.cell(row=year_row, column=col).value
                try:
                    year = int(cell_val)
                    if year <= c.START_YEAR_TECHNO_MODELS or year > c.END_YEAR_TECHNO_MODELS:
                        cols_to_delete.append(col)
                except (TypeError, ValueError):
                    continue

            for col_idx in reversed(cols_to_delete):
                ws.delete_cols(col_idx)

    wb.save(file_path)
    wb.close()

    return {"success": True, "message": f"Model '{model}' created successfully."}

@router.get("/download-technoeconomic-model-files/{model}")
def download_techno_model_files(model: str, background_tasks: BackgroundTasks):
    temp_dir = tempfile.mkdtemp()
    try:
        paths = [
            c.FUEL_MARKET_INFORMATION_PATH,
            f"{c.TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx",
            f"{c.DESIGN_CAPITAL_STRUCTURE_MODEL}{model}.xlsx",
        ]

        found_model_file = None
        for f in os.listdir(c.TECHNOECONOMIC_MODELS_FOLDER):
            if model.lower() in f.lower():
                candidate_path = os.path.join(c.TECHNOECONOMIC_MODELS_FOLDER, f)
                if os.path.isfile(candidate_path):
                    found_model_file = candidate_path
                    break

        if found_model_file:
            paths.append(found_model_file)

        for path in paths:
            if not os.path.exists(path):
                shutil.rmtree(temp_dir)
                return {"error": f"Missing file: {path}"}
            shutil.copy(path, os.path.join(temp_dir, os.path.basename(path)))

        zip_path = os.path.join(tempfile.gettempdir(), f"{model}_technoeconomic_files.zip")
        shutil.make_archive(zip_path.replace(".zip", ""), 'zip', temp_dir)

        background_tasks.add_task(os.remove, zip_path)
        shutil.rmtree(temp_dir)

        return FileResponse(
            zip_path,
            media_type="application/zip",
            filename=f"{model}_technoeconomic_files.zip"
        )
    except Exception as e:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        return {"error": str(e)}

@router.delete("/delete-technoeconomic-model/{model}")
def delete_techno_model(model: str):
    errors = []
    model_lower = model.lower()

    if model_lower == "bau":
        for folder in c.FOLDERS:
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                if os.path.isfile(file_path) and "template" not in filename.lower():
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        errors.append(f"Failed to delete {file_path}: {str(e)}")
    else:
        paths_to_delete = [
            os.path.join(c.TECHNOECONOMIC_MODELS_FOLDER, f)
            for f in os.listdir(c.TECHNOECONOMIC_MODELS_FOLDER)
            if model_lower in f.lower()
        ]

        related_files = [
            f"{c.TECHNOECONOMIC_INPUTS_MODEL}{model}.xlsx",
            f"{c.DESIGN_CAPITAL_STRUCTURE_MODEL}{model}.xlsx",
        ]

        for path in related_files:
            paths_to_delete.append(path)

        for path in paths_to_delete:
            try:
                if os.path.exists(path) and os.path.abspath(path) != os.path.abspath(c.FUEL_MARKET_INFORMATION_PATH):
                    os.remove(path)
            except Exception as e:
                errors.append(f"Failed to delete {path}: {str(e)}")

    try:
        carbon_file = f"{c.CARBON_CREDITS_PATH}"
        sheet = "Carbon Credits"

        df = pd.read_excel(carbon_file, sheet_name=sheet)
        first_col = df.columns[0]
        original_col = df[first_col]

        mask = ~original_col.astype(str).str.lower().str.contains(model.lower())
        df_filtered = df[mask]

        with pd.ExcelWriter(carbon_file, mode='a', if_sheet_exists='replace') as writer:
            df_filtered.to_excel(writer, sheet_name=sheet, index=False)
    except Exception as e:
        errors.append(f"Failed to update Carbon Credits file: {str(e)}")
    if errors:
        raise HTTPException(status_code=500, detail={"errors": errors})

    return {"status": "deleted"}

@router.post("/upload-technoeconomic-model/{name}")
def upload_techno_model(name: str, file: UploadFile = File(...)):
    ext = os.path.splitext(file.filename)[1].lower()

    if ext not in c.VALID_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only Excel files are allowed.")

    contents = file.file.read()
    new_file = load_workbook(filename=BytesIO(contents), data_only=True, read_only=True)

    new_start = new_file["Contents"]["G8"].value
    new_end = new_file["Contents"]["G24"].value

    if name.lower() == "bau":
        reference_path = c.CARBON_CREDITS_PATH
        if not os.path.exists(reference_path):
            shutil.copy(c.CARBON_CREDITS_TEMPLATE_PATH, reference_path)
    else:
        reference_path = f"{c.TECHNOECONOMIC_INPUTS_MODEL}{name}.xlsx"
        if not os.path.exists(reference_path):
            raise HTTPException(status_code=400, detail=f"Reference technoeconomic model not found: {reference_path}")

    ref_start, ref_end = detect_year_range(reference_path)
    if (new_start != ref_start) or (new_end != ref_end):
        raise HTTPException(
            status_code=400,
            detail=f"Year range mismatch: expected {ref_start}-{ref_end}, got {new_start}-{new_end}"
        )

    UPLOADED_MODEL_PATH = os.path.join(c.TECHNOECONOMIC_MODELS_FOLDER, name + ext)
    with open(UPLOADED_MODEL_PATH, "wb") as f:
            f.write(contents)

    e.fill_contents_from_source(UPLOADED_MODEL_PATH, c.CARBON_CREDITS_PATH, name, carbon_flag=True)

    if name.lower() != "bau":
        e.fill_contents_from_source(UPLOADED_MODEL_PATH, reference_path, name, carbon_flag=False)

    return {"status": "uploaded"}
