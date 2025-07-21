from openpyxl import load_workbook

def detect_year_range(path):
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                if isinstance(ws.cell(row=row, column=col).value, str) and ws.cell(row=row, column=col).value.strip().lower() == "baseline":
                    years = []
                    for c in range(col + 1, ws.max_column + 1):
                        try:
                            y = int(ws.cell(row=row, column=c).value)
                            years.append(y)
                        except Exception:
                            continue
                    if years:
                        return min(years) - 1, max(years)
    return None, None
